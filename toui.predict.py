from keras.utils import np_utils
from keras.models import load_model
from PIL import Image
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, roc_curve, roc_auc_score, auc
import numpy as np
import os
import glob
import matplotlib.pyplot as plt
import openpyxl as xl
import pandas as pd
import ezodf

test_dir = '/home/toui/デスクトップ/ori/add_data/DATA7_test/test_50/test/'

excel_dir = "/home/toui/デスクトップ/kekka/kekkax2.xlsx"
book = xl.load_workbook(excel_dir)
# book = ezodf.opendoc(filename=os.path.basename(excel_dir)).sheets[0]
sheet = book["Sheet1"]

classes = ["broken", "correct"]
img_width, img_height = 200, 200

nb_classes = len(classes)

# データセットの作成
x_test = []
y_test = []
test_name = []

result_dir = "/home/toui/デスクトップ/kekka/"

def make_dataset(path, x, y, z):
    for index, name in enumerate(classes):
        dir = path + name
        files = glob.glob(dir + "/*.png")
        for i, file in enumerate(files):
            image = Image.open(file)
            image = image.convert("L").convert("RGB")
            image = image.resize((img_height, img_width))
            data = np.asarray(image)
            x.append(data)
            y.append(index)
            z.append(os.path.basename(file))

    x = np.array(x)
    y = np.array(y)
    # z = np.array(z)

    x = x.astype("float32")
    x = x / 255.0

    y = np_utils.to_categorical(y, nb_classes)

    return x, y, z

x_test, y_test, file_name = make_dataset(test_dir, x_test, y_test, test_name)

#json_string = open(os.path.join(model_dir, model_filename)).read()

#モデルの選択

model = load_model(os.path.join(result_dir + 'model.h5'))

model.load_weights(os.path.join(result_dir, "weight.h5"))

y_preds = model.predict(x_test)
scores = []

for x in y_preds:
    scores.append(x[0])

y_pred_ = np.argmax(y_preds, axis = 1)
y_test_ = np.argmax(y_test, axis = 1)

print("accuracy=",accuracy_score(y_test_, y_pred_))
print(classification_report(y_test_, y_pred_))
print(confusion_matrix(y_test_, y_pred_))

tp, fn, fp, tn = confusion_matrix(y_test_, y_pred_).flatten()
print(tp,fn,fp,tn)
TPR = tp / (tp + fn)
FPR = fp / (fp + tn)
print("TPR="+str(TPR)+" "+"FPR="+str(FPR))

fpr, tpr, thresholds = roc_curve(y_test_, y_pred_,drop_intermediate=False)
auc = auc(fpr, tpr)
plt.plot(fpr, tpr, label='ROC curve (area = %.2f)'%auc)
plt.legend()
plt.title('ROC curve')
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.grid(True)
# plt.plot(fpr, tpr, marker='o')
# plt.xlabel('FPR: False positive rate')
# plt.ylabel('TPR: True positive rate')
# plt.grid(True)
plt.show()
print("auc="+str(roc_auc_score(y_test_, y_pred_)))

#sheetにどの画像があっているか書き出し
index = 3

val = []
ans = []
for i in range(len(file_name)):
    file_name[i] = file_name[i].replace(".png", "").replace("-", "").replace("0","",1)
    val = int(file_name[i]), y_test_[i], y_pred_[i]
    ans.append(val)
ans.sort(key=lambda x: x[0])


for i in range(len(file_name)):
    # sheet.cell(row=index, column=2).value = file_name[i]
    # sheet.cell(row=index, column=3).value = y_test_[i]
    # sheet.cell(row=index, column=5).value = y_pred_[i]

    sheet.cell(row=index, column=2).value = ans[i][0]
    sheet.cell(row=index, column=3).value = ans[i][1]
    sheet.cell(row=index, column=5).value = ans[i][2]

    index += 1

book.save(excel_dir)